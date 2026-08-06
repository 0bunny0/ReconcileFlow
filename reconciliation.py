from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import math
import re
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


class ReconciliationError(ValueError):
    pass


INVALID_ORDER_NUMBERS = {"", "-", "—", "NAN", "NONE", "NULL"}
ZERO_DIFFERENCE_EPSILON = 1e-9

BASIS_CONFIG: dict[str, dict[str, Any]] = {
    "rmb": {
        "label": "人民币口径",
        "short_label": "RMB",
        "document_aliases": ("VAT PRICE", "VAT  PRICE"),
        "delivery_aliases": ("出货运费(RMB)", "出货运费（RMB）"),
        "freight_label": "出货运费(RMB)",
        "system_label": "系统金额(RMB)",
        "document_label": "文控 VAT PRICE",
        "difference_label": "差异(RMB)",
        "number_format": '#,##0.00;[Red]-#,##0.00;–',
    },
    "original": {
        "label": "原币口径",
        "short_label": "原币",
        "document_aliases": ("TP-CPO",),
        "delivery_aliases": ("出货运费(原币)", "出货运费（原币）"),
        "freight_label": "出货运费(原币)",
        "system_label": "系统金额(原币)",
        "document_label": "文控 TP-CPO",
        "difference_label": "差异(原币)",
        "number_format": '#,##0.0000;[Red]-#,##0.0000;–',
    },
}

MODULE_CONFIG: dict[str, dict[str, Any]] = {
    "order": {
        "label": "接单差异核对",
        "short_label": "接单",
        "document_keyword": "接单",
        "amount_source_label": "接单金额明细",
        "freight_source_label": "接单运费明细",
        "document_sheet_label": "文控接单明细",
        "amount_aliases": {
            "rmb": ("接单金额(RMB)", "接单金额（RMB）"),
            "original": ("交易金额",),
        },
        "append_columns": {"rmb": "接单金额(RMB)", "original": "交易金额"},
        "amount_labels": {"rmb": "接单金额(RMB)", "original": "交易金额(原币)"},
        "amount_date_aliases": ("接单时间", "接单日期"),
    },
    "shipment": {
        "label": "出货差异核对",
        "short_label": "出货",
        "document_keyword": "出货",
        "amount_source_label": "出货金额明细",
        "freight_source_label": "出货运费明细",
        "document_sheet_label": "文控出货明细",
        "amount_aliases": {
            "rmb": ("出货金额(RMB)", "出货金额（RMB）", "实际出货金额(RMB)", "实际出货金额（RMB）"),
            "original": ("出货金额", "实际出货金额"),
        },
        "append_columns": {"rmb": "出货金额(RMB)", "original": "出货金额"},
        "amount_labels": {"rmb": "出货金额(RMB)", "original": "出货金额(原币)"},
        "amount_date_aliases": ("出货日期", "实际出货日期", "实际出厂日期"),
    },
}


def basis_config(basis: str) -> dict[str, Any]:
    try:
        return BASIS_CONFIG[basis]
    except KeyError as exc:
        raise ReconciliationError("核对口径必须选择人民币或原币。") from exc


def module_config(module: str) -> dict[str, Any]:
    try:
        return MODULE_CONFIG[module]
    except KeyError as exc:
        raise ReconciliationError("核对模块必须选择接单或出货。") from exc


@dataclass
class Result:
    customer_all: pd.DataFrame
    customer_differences: pd.DataFrame
    line_all: pd.DataFrame
    line_differences: pd.DataFrame
    document_rows: pd.DataFrame
    system_rows: pd.DataFrame
    warnings: list[str]
    statistics: dict[str, Any]
    source_names: dict[str, str]
    tolerance: float
    basis: str
    module: str


def norm_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).replace("（", "(").replace("）", ")").upper()


def is_missing(value: Any) -> bool:
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def norm_id(value: Any) -> str:
    if is_missing(value):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if re.fullmatch(r"[-+]?\d+\.0+", text):
        text = text.split(".", 1)[0]
    return re.sub(r"\s+", "", text).upper()


def norm_currency(value: Any) -> str:
    currency = norm_id(value)
    return "CNY" if currency in {"RMB", "¥", "￥"} else currency


def valid_order(value: str) -> bool:
    return value not in INVALID_ORDER_NUMBERS


def numeric(series: pd.Series, fill_zero: bool = True) -> pd.Series:
    cleaned = series.map(lambda value: None if is_missing(value) else str(value).replace(",", "").strip())
    result = pd.to_numeric(cleaned, errors="coerce").astype(float)
    return result.fillna(0.0) if fill_zero else result


def workbook_bytes(source: Any) -> bytes:
    if isinstance(source, (bytes, bytearray)):
        return bytes(source)
    if isinstance(source, (str, Path)):
        return Path(source).read_bytes()
    if hasattr(source, "seek"):
        source.seek(0)
    data = source.read()
    if hasattr(source, "seek"):
        source.seek(0)
    return data


def read_sheets(source: Any) -> dict[str, pd.DataFrame]:
    if isinstance(source, dict) and all(isinstance(value, pd.DataFrame) for value in source.values()):
        return source
    try:
        return pd.read_excel(BytesIO(workbook_bytes(source)), sheet_name=None, dtype=object, engine="openpyxl")
    except Exception as exc:
        raise ReconciliationError(f"Excel 文件无法读取：{exc}") from exc


def find_column(frame: pd.DataFrame, *aliases: str) -> str | None:
    columns = {norm_header(column): str(column) for column in frame.columns}
    return next((columns[norm_header(alias)] for alias in aliases if norm_header(alias) in columns), None)


def require_columns(frame: pd.DataFrame, specs: dict[str, tuple[str, ...]]) -> dict[str, str] | None:
    resolved = {key: find_column(frame, *aliases) for key, aliases in specs.items()}
    return None if any(value is None for value in resolved.values()) else resolved  # type: ignore[return-value]


def load_document(source: Any, basis: str, module: str = "order") -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    config = basis_config(basis)
    module_info = module_config(module)
    required = {"customer": ("BIL",), "order": ("JO",), "amount": config["document_aliases"]}
    standard_parts: list[pd.DataFrame] = []
    display_parts: list[pd.DataFrame] = []
    warnings: list[str] = []
    for sheet_name, frame in read_sheets(source).items():
        columns = require_columns(frame, required)
        if module_info["document_keyword"] not in str(sheet_name) or not columns:
            continue
        rows = pd.Series(range(2, len(frame) + 2), index=frame.index)
        customers = frame[columns["customer"]].map(norm_id)
        keep = customers.ne("") & frame[columns["amount"]].notna()
        work, rows, customers = frame.loc[keep].copy(), rows.loc[keep], customers.loc[keep]
        orders = work[columns["order"]].map(norm_id)
        amounts = numeric(work[columns["amount"]])
        tp_col, part_col, date_col = find_column(work, "TP-CPO"), find_column(work, "PART-DWG"), find_column(work, "DAY")
        currency_col = find_column(work, "CURR")
        original = numeric(work[tp_col], False) if tp_col else pd.Series(float("nan"), index=work.index)
        currencies = work[currency_col].map(norm_currency) if currency_col else pd.Series("", index=work.index)
        items = work[part_col].map(lambda value: "" if is_missing(value) else str(value)) if part_col else pd.Series("", index=work.index)
        dates = pd.to_datetime(work[date_col], errors="coerce") if date_col else pd.Series(pd.NaT, index=work.index)
        keys = [order if valid_order(order) else f"__DOC__{sheet_name}__{row}" for order, row in zip(orders, rows)]
        methods = ["流水号直接匹配" if valid_order(order) else "未匹配" for order in orders]
        standard = pd.DataFrame({
            "客户代码": customers, "订单流水号": orders, "匹配键": keys,
            "关联流水号": [order if valid_order(order) else "" for order in orders], "匹配方式": methods,
            "文控金额": amounts, "文控原币金额": original, "币种": currencies,
            "文控项目": items, "文控日期": dates,
            "来源表": sheet_name, "来源行号": rows.astype(int),
        })
        detail = work.copy().reset_index(drop=True)
        metadata = pd.DataFrame({
            "客户代码_标准": customers.reset_index(drop=True),
            "订单流水号_标准": orders.reset_index(drop=True),
            "匹配键": keys,
            "关联流水号": [order if valid_order(order) else "" for order in orders],
            "匹配方式": methods,
            "来源表": sheet_name,
            "来源行号": rows.astype(int).reset_index(drop=True),
        })
        detail = pd.concat([metadata, detail], axis=1)
        standard_parts.append(standard.reset_index(drop=True))
        display_parts.append(detail.reset_index(drop=True))
        invalid = sum(not valid_order(order) for order in orders)
        if invalid:
            warnings.append(f"文控子表“{sheet_name}”有 {invalid} 行缺少有效 JO，未能匹配的行会单独列示。")
    if not standard_parts:
        amount_name = "VAT PRICE" if basis == "rmb" else "TP-CPO"
        keyword = module_info["document_keyword"]
        raise ReconciliationError(f"文控登记表中没有找到名称包含“{keyword}”且含 BIL、JO、{amount_name} 的子表。")
    return pd.concat(standard_parts, ignore_index=True), pd.concat(display_parts, ignore_index=True), warnings


def load_system(source: Any, role: str, basis: str, module: str = "order") -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    config = basis_config(basis)
    module_info = module_config(module)
    is_amount = role == "amount"
    source_label = module_info["amount_source_label"] if is_amount else module_info["freight_source_label"]
    amount_aliases = module_info["amount_aliases"][basis] if is_amount else config["delivery_aliases"]
    required = {"customer": ("客户代码",), "order": ("订单流水号",), "amount": amount_aliases}
    standard_parts: list[pd.DataFrame] = []
    display_parts: list[pd.DataFrame] = []
    warnings: list[str] = []
    for sheet_name, frame in read_sheets(source).items():
        columns = require_columns(frame, required)
        if not columns:
            continue
        rows = pd.Series(range(2, len(frame) + 2), index=frame.index)
        customers = frame[columns["customer"]].map(norm_id)
        keep = customers.ne("")
        work, rows, customers = frame.loc[keep].copy(), rows.loc[keep], customers.loc[keep]
        orders = work[columns["order"]].map(norm_id)
        amount = numeric(work[columns["amount"]])
        keys = [order if valid_order(order) else f"__{module}__{role}__{sheet_name}__{row}" for order, row in zip(orders, rows)]
        original_col = None if is_amount else find_column(work, "出货运费(原币)", "出货运费（原币）")
        currency_col = find_column(work, "交易币种", "币种", "CURR")
        date_aliases = module_info["amount_date_aliases"] if is_amount else ("实际出厂日期", "出厂日期", "出货日期", "实际出货日期")
        date_col = find_column(work, *date_aliases)
        original = numeric(work[original_col], False) if original_col else pd.Series(float("nan"), index=work.index)
        currencies = work[currency_col].map(norm_currency) if currency_col else pd.Series("", index=work.index)
        dates = pd.to_datetime(work[date_col], errors="coerce") if date_col else pd.Series(pd.NaT, index=work.index)
        standard = pd.DataFrame({
            "客户代码": customers, "订单流水号": orders, "匹配键": keys,
            "主金额": amount if is_amount else 0.0, "运费": 0.0 if is_amount else amount,
            "系统金额": amount, "运费原币": original, "币种": currencies, "系统日期": dates,
            "数据来源": source_label, "来源表": sheet_name, "来源行号": rows.astype(int),
        })
        detail = work.copy().reset_index(drop=True)
        append_column = module_info["append_columns"][basis]
        detail[append_column] = amount.reset_index(drop=True)
        metadata = pd.DataFrame({
            "客户代码_标准": customers.reset_index(drop=True),
            "订单流水号_标准": orders.reset_index(drop=True),
            "匹配键": keys,
            "数据来源": source_label,
            "来源表": sheet_name,
            "来源行号": rows.astype(int).reset_index(drop=True),
        })
        detail = pd.concat([metadata, detail], axis=1)
        standard_parts.append(standard.reset_index(drop=True))
        display_parts.append(detail.reset_index(drop=True))
        invalid = sum(not valid_order(order) for order in orders)
        if invalid:
            warnings.append(f"{source_label}“{sheet_name}”有 {invalid} 行缺少有效订单流水号，已按原始行单独列示。")
    if not standard_parts:
        raise ReconciliationError(f"{source_label}中没有找到客户代码、订单流水号和{amount_aliases[0]}。")
    return pd.concat(standard_parts, ignore_index=True), pd.concat(display_parts, ignore_index=True), warnings


def auto_match_freight(document: pd.DataFrame, detail: pd.DataFrame, delivery: pd.DataFrame) -> tuple[int, int]:
    candidates = set(document.index[
        document["订单流水号"].map(lambda value: not valid_order(value))
        & document["文控项目"].map(lambda value: "SHIPPINGCOST" in norm_header(value) or "运费" in str(value))
        & document["文控原币金额"].notna()
    ])
    matched = 0
    eligible = 0
    for delivery_index in delivery.sort_values(["客户代码", "来源行号"]).index:
        row = delivery.loc[delivery_index]
        amount = row["运费原币"]
        if is_missing(amount) or not math.isfinite(float(amount)):
            continue
        eligible += 1
        choices = [index for index in candidates if document.at[index, "客户代码"] == row["客户代码"] and math.isclose(float(document.at[index, "文控原币金额"]), float(amount), abs_tol=1e-6, rel_tol=0)]
        if not choices:
            continue
        def rank(index: int) -> tuple[float, int]:
            left, right = document.at[index, "文控日期"], row["系统日期"]
            distance = float("inf") if is_missing(left) or is_missing(right) else abs((pd.Timestamp(left) - pd.Timestamp(right)).total_seconds())
            return distance, int(document.at[index, "来源行号"])
        chosen = min(choices, key=rank)
        key, order = str(row["匹配键"]), str(row["订单流水号"])
        document.loc[chosen, ["匹配键", "关联流水号", "匹配方式"]] = [key, order, "运费自动匹配"]
        mask = detail["来源表"].eq(document.at[chosen, "来源表"]) & detail["来源行号"].eq(document.at[chosen, "来源行号"])
        detail.loc[mask, ["匹配键", "关联流水号", "匹配方式"]] = [key, order, "运费自动匹配"]
        candidates.remove(chosen)
        matched += 1
    return matched, max(eligible - matched, 0)


def status(merge: str, difference: float, tolerance: float) -> str:
    if merge == "left_only": return "仅文控表"
    if merge == "right_only": return "仅系统表"
    return "有差异" if abs(float(difference)) > tolerance else "一致"


def summaries(document: pd.DataFrame, system: pd.DataFrame, tolerance: float) -> tuple[pd.DataFrame, pd.DataFrame]:
    doc_customer = document.groupby("客户代码", as_index=False).agg(文控金额=("文控金额", "sum"), 文控行数=("文控金额", "size"))
    sys_customer = system.groupby("客户代码", as_index=False).agg(主金额=("主金额", "sum"), 运费=("运费", "sum"), 系统金额=("系统金额", "sum"), 系统行数=("系统金额", "size"))
    customers = doc_customer.merge(sys_customer, on="客户代码", how="outer", indicator=True)
    for col in ["文控金额", "文控行数", "主金额", "运费", "系统金额", "系统行数"]: customers[col] = customers[col].fillna(0)
    customers["差异"] = (customers["系统金额"] - customers["文控金额"]).round(6)
    customers["状态"] = [status(m, d, tolerance) for m, d in zip(customers["_merge"], customers["差异"])]
    customers = customers.drop(columns="_merge")[["客户代码", "主金额", "运费", "系统金额", "文控金额", "差异", "状态", "系统行数", "文控行数"]]
    order = {"仅文控表": 0, "仅系统表": 1, "有差异": 2, "一致": 3}
    customers["_sort"] = customers["状态"].map(order)
    customers = customers.sort_values(["_sort", "客户代码"]).drop(columns="_sort").reset_index(drop=True)

    doc_line = document.groupby(["客户代码", "匹配键"], as_index=False).agg(
        文控金额=("文控金额", "sum"), 文控行数=("文控金额", "size"),
        文控流水号=("订单流水号", lambda values: next((value for value in values if value), "")),
        文控来源行=("来源行号", lambda values: ", ".join(str(int(value)) for value in values)),
    )
    sys_line = system.groupby(["客户代码", "匹配键"], as_index=False).agg(
        主金额=("主金额", "sum"), 运费=("运费", "sum"), 系统金额=("系统金额", "sum"), 系统行数=("系统金额", "size"),
        系统流水号=("订单流水号", lambda values: next((value for value in values if value), "")),
        系统来源行=("来源行号", lambda values: ", ".join(str(int(value)) for value in values)),
    )
    lines = doc_line.merge(sys_line, on=["客户代码", "匹配键"], how="outer", indicator=True)
    for col in ["文控金额", "文控行数", "主金额", "运费", "系统金额", "系统行数"]: lines[col] = lines[col].fillna(0)
    for col in ["文控流水号", "文控来源行", "系统流水号", "系统来源行"]: lines[col] = lines[col].fillna("")
    display_orders = []
    for doc_order, sys_order, doc_rows, sys_rows in zip(lines["文控流水号"], lines["系统流水号"], lines["文控来源行"], lines["系统来源行"]):
        value = doc_order if valid_order(doc_order) else sys_order
        if not valid_order(value): value = f"（无有效流水号｜{'文控第'+doc_rows+'行' if doc_rows else '系统第'+sys_rows+'行'}）"
        display_orders.append(value)
    lines["订单流水号"] = display_orders
    lines["差异"] = (lines["系统金额"] - lines["文控金额"]).round(6)
    lines["状态"] = [status(m, d, tolerance) for m, d in zip(lines["_merge"], lines["差异"])]
    lines = lines.drop(columns="_merge")[["客户代码", "订单流水号", "匹配键", "主金额", "运费", "系统金额", "文控金额", "差异", "状态", "系统行数", "文控行数", "系统来源行", "文控来源行"]]
    return customers, lines.sort_values(["客户代码", "状态", "订单流水号"]).reset_index(drop=True)


def reconcile(
    document_source: Any,
    amount_source: Any,
    freight_source: Any,
    *,
    tolerance: float = 1.0,
    basis: str = "rmb",
    module: str = "order",
    source_names: dict[str, str] | None = None,
) -> Result:
    if tolerance < 0: raise ReconciliationError("差异容差不能小于 0。")
    basis_config(basis)
    module_info = module_config(module)
    document, document_rows, warnings = load_document(document_source, basis, module)
    amount, amount_rows, amount_warnings = load_system(amount_source, "amount", basis, module)
    freight, freight_rows, freight_warnings = load_system(freight_source, "freight", basis, module)
    matched, unmatched = auto_match_freight(document, document_rows, freight)
    system = pd.concat([amount, freight], ignore_index=True)
    system_rows = pd.concat([amount_rows, freight_rows], ignore_index=True, sort=False)
    customer_all, line_all = summaries(document, system, tolerance)
    customer_diff = customer_all[
        customer_all["状态"].ne("一致")
        & customer_all["差异"].abs().gt(ZERO_DIFFERENCE_EPSILON)
    ].reset_index(drop=True)
    line_diff = line_all[
        line_all["状态"].ne("一致")
        & line_all["差异"].abs().gt(ZERO_DIFFERENCE_EPSILON)
    ].reset_index(drop=True)
    warnings += amount_warnings + freight_warnings
    if basis == "original":
        currency_rows = pd.concat(
            [document[["客户代码", "币种"]], system[["客户代码", "币种"]]],
            ignore_index=True,
        )
        currency_rows = currency_rows[currency_rows["币种"].ne("")]
        mixed = []
        for customer, group in currency_rows.groupby("客户代码"):
            currencies = sorted(set(group["币种"]))
            if len(currencies) > 1:
                mixed.append(f"{customer}（{'/'.join(currencies)}）")
        if mixed:
            warnings.append("原币模式发现同一客户代码存在多个币种，汇总金额不可直接相加，请分别核对：" + "、".join(mixed))
    if matched: warnings.append(f"文控登记表中 {matched} 条 JO 为空的运费行，已按客户代码、原币运费及最近日期关联到{module_info['freight_source_label']}的订单流水号。")
    if unmatched: warnings.append(f"{module_info['freight_source_label']}中有 {unmatched} 条运费未在文控登记表找到对应运费行，将作为系统侧差异列示。")
    doc_only = customer_all.loc[customer_all["状态"] == "仅文控表", "客户代码"].tolist()
    sys_only = customer_all.loc[customer_all["状态"] == "仅系统表", "客户代码"].tolist()
    if doc_only: warnings.append("仅文控表存在的客户代码：" + "、".join(doc_only))
    if sys_only: warnings.append("仅系统表存在的客户代码：" + "、".join(sys_only))
    stats = {
        "文控有效行": len(document), "金额明细有效行": len(amount), "运费明细有效行": len(freight), "系统拼接行": len(system),
        "差异客户数": len(customer_diff), "差异流水号数": len(line_diff), "自动关联运费行": matched, "未关联运费行": unmatched,
        "系统总额": float(system["系统金额"].sum()), "文控总额": float(document["文控金额"].sum()),
        "总差异": float(system["系统金额"].sum() - document["文控金额"].sum()),
    }
    defaults = {
        "document": "文控登记表.xlsx",
        "amount": f"{module_info['amount_source_label']}.xlsx",
        "freight": f"{module_info['freight_source_label']}.xlsx",
    }
    return Result(
        customer_all, customer_diff, line_all, line_diff, document_rows, system_rows, warnings, stats,
        source_names or defaults, float(tolerance), basis, module,
    )


def reconcile_module_all(
    document_source: Any,
    amount_source: Any,
    freight_source: Any,
    *,
    module: str,
    tolerance: float = 1.0,
    source_names: dict[str, str] | None = None,
) -> dict[str, Result]:
    """Read one module's three workbooks once and cache its RMB/original results."""
    if tolerance < 0:
        raise ReconciliationError("差异容差不能小于 0。")
    module_info = module_config(module)
    sheets = {
        "document": read_sheets(document_source),
        "amount": read_sheets(amount_source),
        "freight": read_sheets(freight_source),
    }
    names = source_names or {
        "document": "文控登记表.xlsx",
        "amount": f"{module_info['amount_source_label']}.xlsx",
        "freight": f"{module_info['freight_source_label']}.xlsx",
    }
    return {
        basis: reconcile(
            sheets["document"], sheets["amount"], sheets["freight"],
            tolerance=tolerance, basis=basis, module=module, source_names=names,
        )
        for basis in ("rmb", "original")
    }


def reconcile_all(
    document_source: Any,
    order_amount_source: Any,
    order_freight_source: Any,
    shipment_amount_source: Any,
    shipment_freight_source: Any,
    *,
    tolerance: float = 1.0,
    source_names: dict[str, str] | None = None,
) -> dict[str, dict[str, Result]]:
    """Read all five workbooks once, then cache both bases for both modules."""
    if tolerance < 0:
        raise ReconciliationError("差异容差不能小于 0。")
    sheets = {
        "document": read_sheets(document_source),
        "order_amount": read_sheets(order_amount_source),
        "order_freight": read_sheets(order_freight_source),
        "shipment_amount": read_sheets(shipment_amount_source),
        "shipment_freight": read_sheets(shipment_freight_source),
    }
    names = source_names or {}
    output: dict[str, dict[str, Result]] = {"order": {}, "shipment": {}}
    for module in ("order", "shipment"):
        amount_key = f"{module}_amount"
        freight_key = f"{module}_freight"
        module_names = {
            "document": names.get("document", "文控登记表.xlsx"),
            "amount": names.get(amount_key, f"{module_config(module)['amount_source_label']}.xlsx"),
            "freight": names.get(freight_key, f"{module_config(module)['freight_source_label']}.xlsx"),
        }
        for basis in ("rmb", "original"):
            output[module][basis] = reconcile(
                sheets["document"], sheets[amount_key], sheets[freight_key],
                tolerance=tolerance, basis=basis, module=module, source_names=module_names,
            )
    return output


def json_value(value: Any) -> Any:
    if is_missing(value): return None
    if isinstance(value, pd.Timestamp): return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, datetime): return value.isoformat(sep=" ")
    if isinstance(value, date): return value.isoformat()
    if hasattr(value, "item"):
        try: return value.item()
        except (ValueError, AttributeError): pass
    return value


def records(frame: pd.DataFrame) -> list[dict[str, Any]]:
    return [{str(key): json_value(value) for key, value in row.items()} for row in frame.to_dict(orient="records")]


def payload(result: Result) -> dict[str, Any]:
    config = basis_config(result.basis)
    module_info = module_config(result.module)
    amount_label = module_info["amount_labels"][result.basis]
    return {
        "customer_differences": records(result.customer_differences), "line_differences": records(result.line_differences),
        "document_rows": records(result.document_rows), "system_rows": records(result.system_rows),
        "warnings": result.warnings, "statistics": result.statistics, "source_names": result.source_names,
        "tolerance": result.tolerance, "basis": result.basis, "module": result.module,
        "basis_label": config["label"],
        "module_label": module_info["label"],
        "labels": {
            "main": amount_label, "order": amount_label, "freight": config["freight_label"],
            "system": config["system_label"], "document": config["document_label"],
            "difference": config["difference_label"],
        },
    }


EXPORT_CUSTOMERS = ["客户代码", "主金额", "运费", "系统金额", "文控金额", "差异", "状态", "系统行数", "文控行数"]
EXPORT_LINES = ["客户代码", "订单流水号", "主金额", "运费", "系统金额", "文控金额", "差异", "状态", "系统行数", "文控行数", "系统来源行", "文控来源行"]


def _excel_value(value: Any) -> Any:
    if is_missing(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if hasattr(value, "item"):
        try:
            return value.item()
        except (ValueError, AttributeError):
            pass
    return value


def _add_sheet(workbook: Workbook, title: str, frame: pd.DataFrame, basis: str, module: str) -> None:
    worksheet = workbook.create_sheet(title)
    frame = frame.copy()
    worksheet.append([str(column) for column in frame.columns])
    for row in frame.itertuples(index=False, name=None):
        worksheet.append([_excel_value(value) for value in row])

    header_fill = PatternFill("solid", fgColor="0B6F69")
    header_font = Font(name="微软雅黑", charset=134, bold=True, color="FFFFFF")
    body_font = Font(name="宋体", charset=134, size=10)
    thin = Side(style="thin", color="D8E2E0")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)
    worksheet.row_dimensions[1].height = 26
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    config = basis_config(basis)
    module_info = module_config(module)
    rmb_amount_headers = {"VAT PRICE", "接单金额(RMB)", "出货金额(RMB)", "实际出货金额(RMB)", "出货运费(RMB)"}
    original_amount_headers = {"TP-CPO", "交易金额", "出货金额", "实际出货金额", "出货运费(原币)"}
    selected_amount_headers = {
        module_info["amount_labels"][basis], config["freight_label"], config["system_label"],
        config["document_label"], config["difference_label"],
    }
    amount_headers = rmb_amount_headers | original_amount_headers | selected_amount_headers | {"单价"}
    date_headers = {"接单时间", "接单日期", "出货日期", "实际出货日期", "实际出厂日期", "DAY"}
    status_column = None
    for column_index, header_cell in enumerate(worksheet[1], 1):
        header = str(header_cell.value)
        if header == "文控 TP-CPO":
            header_cell.alignment = Alignment(horizontal="left", vertical="center")
        if header == "状态":
            status_column = column_index
        for cell in worksheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2):
            for item in cell:
                item.font = body_font
                item.alignment = Alignment(vertical="center")
                if header == "文控 TP-CPO":
                    item.alignment = Alignment(horizontal="left", vertical="center")
                if header in amount_headers and isinstance(item.value, (int, float)):
                    if header in original_amount_headers or (basis == "original" and header in selected_amount_headers):
                        item.number_format = BASIS_CONFIG["original"]["number_format"]
                    else:
                        item.number_format = BASIS_CONFIG["rmb"]["number_format"]
                elif header in date_headers and isinstance(item.value, (datetime, date)):
                    item.number_format = "yyyy-mm-dd"

        values = [str(header)] + ["" if cell.value is None else str(cell.value) for cell in worksheet[get_column_letter(column_index)][1: min(worksheet.max_row, 300)]]
        width = min(max(max((len(value) for value in values), default=8) * 1.15 + 2, 11), 36)
        if header in amount_headers:
            width = max(width, 17)
        if "流水号" in header or header in {"CPO NO", "出货装箱单号", "出货通知单号"}:
            width = max(width, 22)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    fills = {
        "有差异": PatternFill("solid", fgColor="FFF0D2"),
        "仅文控表": PatternFill("solid", fgColor="FDE2E2"),
        "仅系统表": PatternFill("solid", fgColor="E6EAF8"),
        "一致": PatternFill("solid", fgColor="DFF4EA"),
    }
    if status_column:
        for row in range(2, worksheet.max_row + 1):
            status_cell = worksheet.cell(row, status_column)
            if status_cell.value in fills:
                status_cell.fill = fills[status_cell.value]
                status_cell.font = Font(name="微软雅黑", charset=134, bold=True, color="233534")

    if worksheet.max_row >= 2 and worksheet.max_column >= 1:
        table = Table(displayName=f"Table{len(workbook.worksheets):02d}", ref=worksheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        worksheet.add_table(table)


def export_excel(result: Result) -> bytes:
    """Build an auditable reconciliation workbook and return its XLSX bytes."""
    config = basis_config(result.basis)
    module_info = module_config(result.module)
    export_names = {
        "主金额": module_info["amount_labels"][result.basis], "运费": config["freight_label"],
        "系统金额": config["system_label"], "文控金额": config["document_label"],
        "差异": config["difference_label"],
    }
    customer_differences = result.customer_differences.reindex(columns=EXPORT_CUSTOMERS).rename(columns=export_names)
    line_differences = result.line_differences.reindex(columns=EXPORT_LINES).rename(columns=export_names)
    customer_all = result.customer_all.reindex(columns=EXPORT_CUSTOMERS).rename(columns=export_names)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_sheet(workbook, "差异汇总", customer_differences, result.basis, result.module)
    _add_sheet(workbook, "流水号差异", line_differences, result.basis, result.module)
    _add_sheet(workbook, "全部客户汇总", customer_all, result.basis, result.module)
    _add_sheet(workbook, "系统拼接明细", result.system_rows, result.basis, result.module)
    _add_sheet(workbook, module_info["document_sheet_label"], result.document_rows, result.basis, result.module)

    notes = workbook.create_sheet("核对说明")
    if result.basis == "rmb":
        system_rule = f"{module_info['amount_source_label']}的{module_info['amount_labels']['rmb']} + {module_info['freight_source_label']}的出货运费(RMB)"
        document_rule = f"文控登记表中名称包含“{module_info['document_keyword']}”的子表之 VAT PRICE"
        tolerance_unit = "元"
    else:
        system_rule = f"{module_info['amount_source_label']}的{module_info['append_columns']['original']} + {module_info['freight_source_label']}的出货运费(原币)"
        document_rule = f"文控登记表中名称包含“{module_info['document_keyword']}”的子表之 TP-CPO"
        tolerance_unit = "原币单位"
    note_rows = [
        ("项目", "内容"),
        ("核对模块", module_info["label"]),
        ("核对口径", f"{config['label']}（{config['short_label']}）"),
        ("系统金额", system_rule),
        ("文控金额", document_rule),
        ("差异公式", "系统金额 - 文控金额"),
        ("差异容差", f"绝对差异不超过 {result.tolerance:,.4f} {tolerance_unit}视为一致"),
        ("运费匹配", f"文控 JO 为空且项目为 shipping cost/运费时，按客户代码、原币金额和最近日期关联{module_info['freight_source_label']}流水号；不改写原始 JO。"),
        ("文控文件", result.source_names.get("document", "")),
        ("金额明细文件", result.source_names.get("amount", "")),
        ("运费明细文件", result.source_names.get("freight", "")),
    ]
    for row in note_rows:
        notes.append(row)
    notes.append(("", ""))
    notes.append(("处理统计", ""))
    for key, value in result.statistics.items():
        notes.append((key, _excel_value(value)))
    notes.append(("", ""))
    notes.append(("提示", ""))
    for warning in result.warnings:
        notes.append(("", warning))

    notes.column_dimensions["A"].width = 20
    notes.column_dimensions["B"].width = 100
    notes.freeze_panes = "A2"
    for cell in notes[1]:
        cell.fill = PatternFill("solid", fgColor="0B6F69")
        cell.font = Font(name="微软雅黑", charset=134, bold=True, color="FFFFFF")
    for row in notes.iter_rows():
        for cell in row:
            cell.font = Font(
                name="宋体", charset=134, size=10,
                bold=cell.row == 1 or notes.cell(cell.row, 1).value in {"处理统计", "提示"},
            )
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    notes.row_dimensions[1].height = 26

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()

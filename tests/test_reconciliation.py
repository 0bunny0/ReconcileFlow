from io import BytesIO
from pathlib import Path
from http.client import HTTPConnection
from http.server import ThreadingHTTPServer
import json
import sys
import threading
import unittest

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from reconciliation import export_excel, payload, reconcile, reconcile_all  # noqa: E402
from app import Handler  # noqa: E402


def book(sheets: dict[str, pd.DataFrame]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name, index=False)
    return output.getvalue()


def multipart(fields: dict[str, tuple[str, bytes] | str]) -> tuple[bytes, str]:
    boundary = "----PoowardReconciliationTest"
    chunks: list[bytes] = []
    for name, value in fields.items():
        chunks.append(f"--{boundary}\r\n".encode())
        if isinstance(value, tuple):
            filename, content = value
            chunks.append(f'Content-Disposition: form-data; name="{name}"; filename="{filename}"\r\n'.encode())
            chunks.append(b"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n")
            chunks.append(content)
            chunks.append(b"\r\n")
        else:
            chunks.append(f'Content-Disposition: form-data; name="{name}"\r\n\r\n{value}\r\n'.encode())
    chunks.append(f"--{boundary}--\r\n".encode())
    return b"".join(chunks), f"multipart/form-data; boundary={boundary}"


class ReconciliationTests(unittest.TestCase):
    def setUp(self) -> None:
        document_order = pd.DataFrame([
            {"DAY": "2026-07-01", "BIL": "CB001", "JO": "BD001", "PART-DWG": "PART A", "CURR": "USD", "TP-CPO": 14.2, "VAT  PRICE": 100},
            {"DAY": "2026-07-02", "BIL": "CB001", "JO": "-", "PART-DWG": "DHL shipping cost", "CURR": "USD", "TP-CPO": 10, "VAT  PRICE": 70},
            {"DAY": "2026-07-03", "BIL": "CB002", "JO": "BD003", "PART-DWG": "PART B", "CURR": "USD", "TP-CPO": 27.1, "VAT  PRICE": 190},
            {"DAY": "2026-07-04", "BIL": "CBZERO", "JO": "ZERO001", "PART-DWG": "ZERO PART", "CURR": "USD", "TP-CPO": 0, "VAT  PRICE": 0},
        ])
        document_shipment = pd.DataFrame([
            {"DAY": "2026-07-05", "BIL": "CB001", "JO": "SD001", "PART-DWG": "PART A", "CURR": "USD", "TP-CPO": 20, "VAT  PRICE": 140, "INVOICE": "INV-01"},
            {"DAY": "2026-07-06", "BIL": "CB001", "JO": "-", "PART-DWG": "shipping cost", "CURR": "USD", "TP-CPO": 5, "VAT  PRICE": 35, "INVOICE": "INV-01"},
            {"DAY": "2026-07-07", "BIL": "CB003", "JO": "SD003", "PART-DWG": "PART C", "CURR": "USD", "TP-CPO": 50, "VAT  PRICE": 350, "INVOICE": "INV-02"},
        ])
        order_amount = pd.DataFrame([
            {"接单时间": "2026-07-01", "客户代码": "CB001", "订单流水号": "BD001", "交易金额": 14.2, "接单金额(RMB)": 100, "自定义订单字段": "保留A"},
            {"接单时间": "2026-07-03", "客户代码": "CB002", "订单流水号": "BD003", "交易金额": 28.1, "接单金额(RMB)": 200, "自定义订单字段": "保留B"},
        ])
        order_freight = pd.DataFrame([
            {"实际出厂日期": "2026-07-02", "客户代码": "CB001", "订单流水号": "BD002", "出货运费(原币)": 10, "出货运费(RMB)": 70},
        ])
        shipment_amount = pd.DataFrame([
            {"出货日期": "2026-07-05", "客户代码": "CB001", "订单流水号": "SD001", "出货金额": 20, "出货金额(RMB)": 140, "交易币种": "USD", "装箱单号": "PL-01"},
            {"出货日期": "2026-07-07", "客户代码": "CB003", "订单流水号": "SD003", "出货金额": 52, "出货金额(RMB)": 360, "交易币种": "USD", "装箱单号": "PL-02"},
        ])
        shipment_freight = pd.DataFrame([
            {"实际出厂日期": "2026-07-06", "客户代码": "CB001", "订单流水号": "SD002", "出货运费(原币)": 5, "出货运费(RMB)": 35, "出货装箱单号": "PL-01"},
        ])
        self.document_bytes = book({"7月份接单明细": document_order, "7月份出货明细": document_shipment})
        self.order_amount_bytes = book({"Sheet1": order_amount})
        self.order_freight_bytes = book({"Sheet1": order_freight})
        self.shipment_amount_bytes = book({"Shipping List": shipment_amount})
        self.shipment_freight_bytes = book({"Shipped Delivery Fee": shipment_freight})
        self.order_result = reconcile(self.document_bytes, self.order_amount_bytes, self.order_freight_bytes, tolerance=1)

    def test_order_customer_and_freight_matching(self) -> None:
        rows = self.order_result.customer_all.set_index("客户代码")
        self.assertEqual(rows.at["CB001", "状态"], "一致")
        self.assertAlmostEqual(rows.at["CB001", "系统金额"], 170)
        self.assertAlmostEqual(rows.at["CB001", "文控金额"], 170)
        self.assertAlmostEqual(rows.at["CB002", "差异"], 10)
        self.assertEqual(self.order_result.statistics["自动关联运费行"], 1)
        line = self.order_result.line_all[self.order_result.line_all["订单流水号"] == "BD002"].iloc[0]
        self.assertEqual(line["状态"], "一致")

    def test_zero_difference_rows_are_hidden_from_both_difference_views(self) -> None:
        self.assertIn("CBZERO", set(self.order_result.customer_all["客户代码"]))
        self.assertNotIn("CBZERO", set(self.order_result.customer_differences["客户代码"]))
        self.assertNotIn("ZERO001", set(self.order_result.line_differences["订单流水号"]))

    def test_payload_preserves_raw_fields_and_appends_freight(self) -> None:
        data = payload(self.order_result)
        self.assertTrue(data["document_rows"])
        self.assertTrue(data["system_rows"])
        auto = [row for row in data["document_rows"] if row["匹配方式"] == "运费自动匹配"]
        self.assertEqual(auto[0]["关联流水号"], "BD002")
        freight = [row for row in data["system_rows"] if row["数据来源"] == "接单运费明细"]
        self.assertEqual(freight[0]["接单金额(RMB)"], 70)
        amount = [row for row in data["system_rows"] if row["数据来源"] == "接单金额明细"]
        self.assertEqual(amount[0]["自定义订单字段"], "保留A")

    def test_both_modules_and_both_bases_are_cached(self) -> None:
        results = reconcile_all(
            self.document_bytes, self.order_amount_bytes, self.order_freight_bytes,
            self.shipment_amount_bytes, self.shipment_freight_bytes, tolerance=1,
        )
        self.assertEqual(set(results), {"order", "shipment"})
        self.assertEqual(set(results["shipment"]), {"rmb", "original"})
        shipment_rmb = results["shipment"]["rmb"]
        shipment_original = results["shipment"]["original"]
        rmb_rows = shipment_rmb.customer_all.set_index("客户代码")
        original_rows = shipment_original.customer_all.set_index("客户代码")
        self.assertEqual(rmb_rows.at["CB001", "状态"], "一致")
        self.assertAlmostEqual(rmb_rows.at["CB003", "差异"], 10)
        self.assertAlmostEqual(original_rows.at["CB003", "差异"], 2)
        data = payload(shipment_rmb)
        self.assertEqual(data["module"], "shipment")
        self.assertEqual(data["labels"]["main"], "出货金额(RMB)")
        freight = [row for row in data["system_rows"] if row["数据来源"] == "出货运费明细"]
        self.assertEqual(freight[0]["出货金额(RMB)"], 35)
        self.assertEqual(freight[0]["出货装箱单号"], "PL-01")

    def test_excel_exports_have_module_specific_document_sheet(self) -> None:
        order_book = load_workbook(BytesIO(export_excel(self.order_result)), read_only=True)
        self.assertEqual(order_book.sheetnames, ["差异汇总", "流水号差异", "全部客户汇总", "系统拼接明细", "文控接单明细", "核对说明"])
        shipment_result = reconcile(
            self.document_bytes, self.shipment_amount_bytes, self.shipment_freight_bytes,
            tolerance=1, module="shipment",
        )
        shipment_book = load_workbook(BytesIO(export_excel(shipment_result)), read_only=True)
        self.assertIn("文控出货明细", shipment_book.sheetnames)
        headers = [cell.value for cell in next(shipment_book["差异汇总"].iter_rows(min_row=1, max_row=1))]
        self.assertIn("出货金额(RMB)", headers)

    def test_original_currency_basis(self) -> None:
        result = reconcile(
            self.document_bytes, self.order_amount_bytes, self.order_freight_bytes,
            tolerance=0.01, basis="original",
        )
        rows = result.customer_all.set_index("客户代码")
        self.assertEqual(rows.at["CB001", "状态"], "一致")
        self.assertAlmostEqual(rows.at["CB001", "系统金额"], 24.2)
        self.assertAlmostEqual(rows.at["CB001", "文控金额"], 24.2)
        self.assertAlmostEqual(rows.at["CB002", "差异"], 1.0)
        data = payload(result)
        self.assertEqual(data["basis"], "original")
        self.assertEqual(data["labels"]["document"], "文控 TP-CPO")
        freight = [row for row in data["system_rows"] if row["数据来源"] == "接单运费明细"]
        self.assertEqual(freight[0]["交易金额"], 10)
        workbook = load_workbook(BytesIO(export_excel(result)), read_only=False)
        sheet = workbook["差异汇总"]
        headers = [cell.value for cell in sheet[1]]
        document_column = headers.index("文控 TP-CPO") + 1
        self.assertEqual(sheet.cell(1, document_column).alignment.horizontal, "left")
        self.assertEqual(sheet.cell(2, document_column).alignment.horizontal, "left")

    def test_original_currency_warns_when_customer_has_multiple_currencies(self) -> None:
        document = book({"接单明细": pd.DataFrame([
            {"DAY": "2026-07-01", "BIL": "CBMIX", "JO": "M001", "PART-DWG": "A", "CURR": "USD", "TP-CPO": 10, "VAT PRICE": 70},
            {"DAY": "2026-07-02", "BIL": "CBMIX", "JO": "M002", "PART-DWG": "B", "CURR": "EUR", "TP-CPO": 20, "VAT PRICE": 150},
        ])})
        amount = book({"Sheet": pd.DataFrame([
            {"客户代码": "CBMIX", "订单流水号": "M001", "交易金额": 10, "接单金额(RMB)": 70, "交易币种": "USD"},
            {"客户代码": "CBMIX", "订单流水号": "M002", "交易金额": 20, "接单金额(RMB)": 150, "交易币种": "EUR"},
        ])})
        freight = book({"Sheet": pd.DataFrame([
            {"客户代码": "CBMIX", "订单流水号": "MFEE", "出货运费(原币)": 0, "出货运费(RMB)": 0},
        ])})
        result = reconcile(document, amount, freight, basis="original")
        self.assertTrue(any("多个币种" in warning and "CBMIX" in warning for warning in result.warnings))

    def test_api_accepts_each_module_independently_and_caches_two_bases(self) -> None:
        server = ThreadingHTTPServer(("127.0.0.1", 0), Handler)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        try:
            cases = {
                "order": {
                    "order_amount": ("order-amount.xlsx", self.order_amount_bytes),
                    "order_freight": ("order-freight.xlsx", self.order_freight_bytes),
                },
                "shipment": {
                    "shipment_amount": ("shipment-amount.xlsx", self.shipment_amount_bytes),
                    "shipment_freight": ("shipment-freight.xlsx", self.shipment_freight_bytes),
                },
            }
            for module, module_fields in cases.items():
                with self.subTest(module=module):
                    body, content_type = multipart({
                        "module": module,
                        "document": ("document.xlsx", self.document_bytes),
                        **module_fields,
                        "tolerance": "1.00",
                    })
                    connection = HTTPConnection("127.0.0.1", server.server_port, timeout=30)
                    connection.request("POST", "/api/analyze", body=body, headers={"Content-Type": content_type, "Content-Length": str(len(body))})
                    response = connection.getresponse()
                    data = json.loads(response.read().decode("utf-8"))
                    self.assertEqual(response.status, 200, data)
                    self.assertEqual(set(data["modules"]), {module})
                    self.assertEqual(set(data["modules"][module]), {"rmb", "original"})
                    download_url = data["modules"][module]["rmb"]["download_url"]
                    connection.request("GET", download_url)
                    download = connection.getresponse()
                    exported = download.read()
                    self.assertEqual(download.status, 200)
                    self.assertTrue(exported.startswith(b"PK"))
                    connection.close()
        finally:
            server.shutdown()
            server.server_close()
            thread.join(timeout=5)


if __name__ == "__main__":
    unittest.main()
